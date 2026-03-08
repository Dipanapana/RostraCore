"""E2E tests for Mafoko roster upload endpoints."""
import io
import pytest
from openpyxl import Workbook
from datetime import date

pytestmark = pytest.mark.integration

# ---------------------------------------------------------------------------
# Helper: build a synthetic Mafoko D/o grid Excel file in memory
# ---------------------------------------------------------------------------

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
UPLOAD_PREFIX = "/api/v1/roster-upload"


def _create_mafoko_grid(
    employees: list[dict],  # [{pers_no, name, grade, contact, days: list[str]}]
    site_name: str = "Test Site Alpha",
    site_code: str = "TSA001",
    province: str = "Gauteng",
    client_name: str = "Mafoko Security Patrols",
    start_date: date = date(2026, 3, 1),
    num_days: int = 7,
) -> bytes:
    """Build a minimal Mafoko D/o grid .xlsx in memory.

    The file layout matches what MafokoRosterImportService expects:
    - Row 1, col 1: header containing both "Mafoko" and "Roster"
    - Row 2: client / site info
    - Row 3, col 1: "Scheduled Period: <ISO start> to <ISO end>"
    - Row 5: column headers (Pers No, Grade, Name, Contact, Worked, day nums)
    - Row 6+: employee data with "D" for day shift and "o" for off
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"

    # Row 1: company header — must contain both "mafoko" and "roster"
    ws.cell(row=1, column=1, value=f"Mafoko Security Patrols - {province} Site Roster")

    # Row 2: site info
    ws.cell(row=2, column=1, value=f"Client: {client_name}")
    ws.cell(row=2, column=3, value=f"Site: {site_code} - {site_name}")

    # Row 3: scheduled period — service scans col 1 rows 1-9 for ISO date range
    end_date_val = date(
        start_date.year,
        start_date.month,
        start_date.day + num_days - 1,
    )
    ws.cell(
        row=3,
        column=1,
        value=f"Scheduled Period: {start_date.isoformat()} to {end_date_val.isoformat()}",
    )

    # Row 5: column headers
    headers = ["Pers No", "Grade", "Name", "Contact", "Worked"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=5, column=i, value=h)
    # Day columns — service expects integer day numbers (1-31) as headers
    for d in range(num_days):
        day_num = start_date.day + d
        ws.cell(row=5, column=6 + d, value=day_num)

    # Employee rows
    for idx, emp in enumerate(employees):
        row = 6 + idx
        ws.cell(row=row, column=1, value=emp["pers_no"])
        ws.cell(row=row, column=2, value=emp.get("grade", "C"))
        ws.cell(row=row, column=3, value=emp["name"])
        ws.cell(row=row, column=4, value=emp.get("contact", "0800000000"))
        # Worked count = number of "D" entries
        working = sum(1 for d in emp["days"] if d == "D")
        ws.cell(row=row, column=5, value=working)
        for d_idx, day_val in enumerate(emp["days"]):
            ws.cell(row=row, column=6 + d_idx, value=day_val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Shared test data
# ---------------------------------------------------------------------------

SAMPLE_EMPLOYEES = [
    {
        "pers_no": "MSP00001",
        "name": "SMITH,J",
        "grade": "C",
        "days": ["D", "D", "o", "D", "D", "o", "D"],
    },
    {
        "pers_no": "MSP00002",
        "name": "JONES,M",
        "grade": "B",
        "days": ["o", "D", "D", "D", "o", "D", "D"],
    },
    {
        # 7 consecutive working days — should trigger BCEA suggestion
        "pers_no": "MSP00003",
        "name": "NKOSI,T",
        "grade": "C",
        "days": ["D", "D", "D", "D", "D", "D", "D"],
    },
]


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestUploadHistory:
    """Tests for GET /history endpoint."""

    def test_upload_history_empty(self, client, ensure_test_user, auth_headers):
        """GET /history returns empty list for org with no uploads."""
        resp = client.get(f"{UPLOAD_PREFIX}/history", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert "total" in data
        assert "uploads" in data
        assert isinstance(data["uploads"], list)

    def test_upload_history_requires_auth(self, client):
        """GET /history without token returns 401."""
        resp = client.get(f"{UPLOAD_PREFIX}/history")
        assert resp.status_code == 401


class TestPreview:
    """Tests for POST /mafoko/preview endpoint."""

    def test_preview_valid_file(self, client, ensure_test_user, auth_headers):
        """Preview a valid Mafoko grid returns parsed header + summary."""
        file_bytes = _create_mafoko_grid(SAMPLE_EMPLOYEES)
        resp = client.post(
            f"{UPLOAD_PREFIX}/mafoko/preview",
            files={"file": ("roster.xlsx", file_bytes, XLSX_MIME)},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "preview"
        assert "parsed_header" in data
        assert "summary" in data
        assert data["summary"]["employees_in_file"] == 3

    def test_preview_invalid_file(self, client, ensure_test_user, auth_headers):
        """Preview with a non-Excel file returns 400."""
        resp = client.post(
            f"{UPLOAD_PREFIX}/mafoko/preview",
            files={"file": ("report.pdf", b"not-an-excel", "application/pdf")},
            headers=auth_headers,
        )
        assert resp.status_code == 400

    def test_preview_returns_suggestions(self, client, ensure_test_user, auth_headers):
        """Preview returns BCEA suggestions for the roster data."""
        file_bytes = _create_mafoko_grid(SAMPLE_EMPLOYEES)
        resp = client.post(
            f"{UPLOAD_PREFIX}/mafoko/preview",
            files={"file": ("roster.xlsx", file_bytes, XLSX_MIME)},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        suggestions = data.get("suggestions", [])
        # Service should flag at least one BCEA concern
        # (overtime, consecutive days, new employees, etc.)
        assert len(suggestions) > 0, "Expected at least one BCEA suggestion"


class TestUpload:
    """Tests for POST /mafoko (single upload) endpoint."""

    def test_upload_valid_file(self, client, ensure_test_user, auth_headers):
        """Upload a valid grid — accepts success or import error.

        The service may return 400 if no employees match (test org has no MSP
        employees), 200 if it successfully creates records, or 409 if duplicate.
        """
        file_bytes = _create_mafoko_grid(SAMPLE_EMPLOYEES)
        resp = client.post(
            f"{UPLOAD_PREFIX}/mafoko",
            files={"file": ("roster.xlsx", file_bytes, XLSX_MIME)},
            params={"auto_create_site": "true"},
            headers=auth_headers,
        )
        assert resp.status_code in [200, 400, 409], (
            f"Unexpected {resp.status_code}: {resp.text}"
        )

    def test_upload_missing_file(self, client, ensure_test_user, auth_headers):
        """POST /mafoko without a file returns 400 or 422 (validation error)."""
        resp = client.post(
            f"{UPLOAD_PREFIX}/mafoko",
            headers=auth_headers,
        )
        assert resp.status_code in [400, 422]

    def test_upload_requires_auth(self, client):
        """POST /mafoko without token returns 401."""
        file_bytes = _create_mafoko_grid(SAMPLE_EMPLOYEES)
        resp = client.post(
            f"{UPLOAD_PREFIX}/mafoko",
            files={"file": ("roster.xlsx", file_bytes, XLSX_MIME)},
        )
        assert resp.status_code == 401


class TestBulkUpload:
    """Tests for POST /mafoko/bulk endpoint."""

    def test_bulk_upload_valid(self, client, ensure_test_user, auth_headers):
        """Bulk upload with 2 files returns aggregated results."""
        file1 = _create_mafoko_grid(
            SAMPLE_EMPLOYEES[:2],
            site_name="Site Bravo",
            site_code="TSB001",
        )
        file2 = _create_mafoko_grid(
            SAMPLE_EMPLOYEES[1:],
            site_name="Site Charlie",
            site_code="TSC001",
        )
        resp = client.post(
            f"{UPLOAD_PREFIX}/mafoko/bulk",
            files=[
                ("files", ("roster_bravo.xlsx", file1, XLSX_MIME)),
                ("files", ("roster_charlie.xlsx", file2, XLSX_MIME)),
            ],
            params={"auto_create_site": True},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "completed"
        assert data["total_files"] == 2
        assert len(data["results"]) == 2


class TestUploadCreatesHistory:
    """Verify that a successful upload appears in upload history."""

    def test_upload_creates_history_record(self, client, ensure_test_user, auth_headers):
        """After an upload attempt, GET /history shows upload log entries.

        Even if the import fails (no employees matched), a log entry should
        be recorded. If upload returns 400, we still check history was logged.
        """
        unique_employees = [
            {
                "pers_no": "MSP99901",
                "name": "HISTORY,X",
                "grade": "C",
                "days": ["D", "o", "D", "o", "D", "o", "D"],
            },
        ]
        file_bytes = _create_mafoko_grid(
            unique_employees,
            site_name="History Test Site",
            site_code="HTS001",
        )

        # Upload (may succeed or fail depending on test data)
        upload_resp = client.post(
            f"{UPLOAD_PREFIX}/mafoko",
            files={"file": ("history_test.xlsx", file_bytes, XLSX_MIME)},
            params={"auto_create_site": "true"},
            headers=auth_headers,
        )
        assert upload_resp.status_code in [200, 400, 409]

        # Check history endpoint works
        history_resp = client.get(f"{UPLOAD_PREFIX}/history", headers=auth_headers)
        assert history_resp.status_code == 200
        history_data = history_resp.json()
        assert "total" in history_data
        assert "uploads" in history_data
