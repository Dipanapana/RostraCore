"""
Roster Template Download & Upload endpoints.

- GET /download: Download empty Excel template for a site/client/period
- POST /upload: Upload filled roster, validate, and create Roster + ShiftAssignments
"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime, timedelta
import io
import logging
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.api.deps import get_current_user
from app.auth.security import get_current_org_id
from app.models.user import User
from app.models.employee import Employee, EmployeeStatus
from app.models.client import Client
from app.models.site import Site
from app.models.shift import Shift, ShiftStatus
from app.models.shift_assignment import ShiftAssignment, AssignmentStatus
from app.models.roster import Roster

logger = logging.getLogger(__name__)

router = APIRouter()

# ── Styles ────────────────────────────────────────────────────────────────────
BLUE_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1E40AF")
META_FONT = Font(name="Calibri", size=10, color="4B5563")
DATA_FONT = Font(name="Calibri", size=10, color="1F2937")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


@router.get("/download")
async def download_roster_template(
    period_start: str = Query(..., description="Start date (YYYY-MM-DD)"),
    period_end: str = Query(..., description="End date (YYYY-MM-DD)"),
    site_id: Optional[int] = Query(None, description="Filter by site ID"),
    client_id: Optional[int] = Query(None, description="Filter by client ID"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    org_id: int = Depends(get_current_org_id),
):
    """
    Download an empty roster template as an Excel file.
    Pre-fills site/client info and generates date rows for the period.
    """
    try:
        start_date = datetime.strptime(period_start, "%Y-%m-%d").date()
        end_date = datetime.strptime(period_end, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    if end_date < start_date:
        raise HTTPException(status_code=400, detail="end_date must be after start_date.")

    if (end_date - start_date).days > 90:
        raise HTTPException(status_code=400, detail="Period cannot exceed 90 days.")

    # Look up site/client info
    site_name = ""
    client_name = ""
    if site_id:
        site = db.query(Site).filter(Site.site_id == site_id, Site.org_id == org_id).first()
        if site:
            site_name = site.site_name or ""
            client_name = site.client_name or ""
    if client_id and not client_name:
        client = db.query(Client).filter(Client.client_id == client_id, Client.org_id == org_id).first()
        if client:
            client_name = client.client_name or ""

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster Template"

    # ── Header section ────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    ws["A1"] = "RostraCore — Roster Upload Template"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Metadata rows
    meta_rows = [
        ("Client:", client_name or "(enter client name)"),
        ("Site:", site_name or "(enter site name)"),
        ("Period:", f"{start_date.strftime('%d %b %Y')} — {end_date.strftime('%d %b %Y')}"),
        ("Generated:", datetime.now().strftime("%d %b %Y %H:%M")),
    ]
    for i, (label, value) in enumerate(meta_rows, 2):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(name="Calibri", size=10, bold=True, color="4B5563")
        ws[f"B{i}"] = value
        ws[f"B{i}"].font = META_FONT

    # Instructions row
    ws.merge_cells("A7:I7")
    ws["A7"] = "Fill in guard assignments below. Employee ID or full name is required. System will validate on upload."
    ws["A7"].font = Font(name="Calibri", size=9, italic=True, color="6B7280")

    # ── Column headers ────────────────────────────────────────────────────
    headers = [
        ("A", "Date", 14),
        ("B", "Day", 10),
        ("C", "Shift Start", 14),
        ("D", "Shift End", 14),
        ("E", "Guard Name", 28),
        ("F", "Employee ID", 14),
        ("G", "PSIRA Number", 16),
        ("H", "Grade", 12),
        ("I", "Notes", 30),
    ]

    header_row = 9
    for col_letter, title, width in headers:
        cell = ws[f"{col_letter}{header_row}"]
        cell.value = title
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[header_row].height = 22

    # ── Pre-fill date rows ────────────────────────────────────────────────
    row = header_row + 1
    current_date = start_date
    day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    while current_date <= end_date:
        day_name = day_names[current_date.weekday()]

        # Two shifts per day (day and night) as starting point
        for shift_type, s_start, s_end in [("Day", "06:00", "18:00"), ("Night", "18:00", "06:00")]:
            ws[f"A{row}"] = current_date.strftime("%Y-%m-%d")
            ws[f"A{row}"].font = DATA_FONT
            ws[f"A{row}"].alignment = Alignment(horizontal="center")

            ws[f"B{row}"] = day_name
            ws[f"B{row}"].font = DATA_FONT
            ws[f"B{row}"].alignment = Alignment(horizontal="center")

            ws[f"C{row}"] = s_start
            ws[f"C{row}"].font = DATA_FONT
            ws[f"C{row}"].alignment = Alignment(horizontal="center")

            ws[f"D{row}"] = s_end
            ws[f"D{row}"].font = DATA_FONT
            ws[f"D{row}"].alignment = Alignment(horizontal="center")

            # Light blue fill for weekend rows
            if current_date.weekday() >= 5:
                for col in "ABCDEFGHI":
                    ws[f"{col}{row}"].fill = LIGHT_BLUE_FILL

            # Apply borders
            for col in "ABCDEFGHI":
                ws[f"{col}{row}"].border = THIN_BORDER

            row += 1

        current_date += timedelta(days=1)

    # Freeze header row
    ws.freeze_panes = f"A{header_row + 1}"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Roster_Template"
    if client_name and client_name != "(enter client name)":
        filename += f"_{client_name.replace(' ', '_')}"
    if site_name and site_name != "(enter site name)":
        filename += f"_{site_name.replace(' ', '_')}"
    filename += f"_{period_start}_to_{period_end}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/upload")
async def upload_roster_template(
    file: UploadFile = File(..., description="Filled roster Excel template"),
    site_id: Optional[int] = Query(None, description="Site ID for the roster"),
    client_id: Optional[int] = Query(None, description="Client ID for the roster"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    org_id: int = Depends(get_current_org_id),
):
    """
    Upload a filled roster template. Validates rows, creates Roster + Shifts + ShiftAssignments.
    Returns an import summary with created count, warnings, and errors.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx)")

    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {str(e)}")

    # Find header row (look for "Date" in column A)
    header_row = None
    for r in range(1, min(20, ws.max_row + 1)):
        val = ws[f"A{r}"].value
        if val and str(val).strip().lower() == "date":
            header_row = r
            break

    if header_row is None:
        raise HTTPException(status_code=400, detail="Could not find header row. Ensure 'Date' header exists in column A.")

    # Parse metadata for period dates
    period_start = None
    period_end = None

    # Read data rows
    assignments_created = 0
    warnings = []
    errors = []
    rows_processed = 0

    # Pre-load employees for this org (for lookups)
    employees = db.query(Employee).filter(
        Employee.org_id == org_id,
        Employee.status == EmployeeStatus.ACTIVE,
    ).all()
    emp_by_id = {e.employee_id: e for e in employees}
    emp_by_name = {f"{e.first_name} {e.last_name}".lower(): e for e in employees}

    # Determine site
    site = None
    if site_id:
        site = db.query(Site).filter(Site.site_id == site_id, Site.org_id == org_id).first()
    if not site:
        # Try to find site from metadata in the spreadsheet
        for r in range(2, min(8, ws.max_row + 1)):
            label = ws[f"A{r}"].value
            if label and "site" in str(label).lower():
                site_val = ws[f"B{r}"].value
                if site_val:
                    site = db.query(Site).filter(
                        Site.org_id == org_id,
                        Site.site_name.ilike(f"%{site_val}%"),
                    ).first()
                break

    # Create Roster record
    now = datetime.utcnow()
    roster_code = f"R-UPLOAD-{now.strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"

    # Collect all date values to determine period
    all_dates = []

    # First pass: collect dates
    for r in range(header_row + 1, ws.max_row + 1):
        date_val = ws[f"A{r}"].value
        if not date_val:
            continue
        try:
            if isinstance(date_val, datetime):
                all_dates.append(date_val.date())
            elif isinstance(date_val, str):
                all_dates.append(datetime.strptime(date_val.strip(), "%Y-%m-%d").date())
        except (ValueError, AttributeError):
            pass

    if not all_dates:
        raise HTTPException(status_code=400, detail="No valid dates found in the template.")

    period_start = min(all_dates)
    period_end = max(all_dates)

    roster = Roster(
        org_id=org_id,
        roster_code=roster_code,
        name=f"Uploaded Roster {period_start.strftime('%d %b')} - {period_end.strftime('%d %b %Y')}",
        start_date=datetime.combine(period_start, datetime.min.time()),
        end_date=datetime.combine(period_end, datetime.max.time()),
        client_id=client_id,
        status="published",
        total_shifts=0,
        assigned_shifts=0,
        unassigned_shifts=0,
        total_cost=0.0,
        regular_pay_cost=0.0,
        overtime_cost=0.0,
        premium_cost=0.0,
        algorithm_used="manual_upload",
        created_by=current_user.user_id,
        published_at=now,
        published_by=current_user.user_id,
    )
    db.add(roster)
    db.flush()  # Get roster_id

    # Second pass: process rows
    total_cost = 0.0
    shifts_created = 0

    for r in range(header_row + 1, ws.max_row + 1):
        date_val = ws[f"A{r}"].value
        if not date_val:
            continue

        rows_processed += 1

        # Parse date
        try:
            if isinstance(date_val, datetime):
                shift_date = date_val.date()
            elif isinstance(date_val, str):
                shift_date = datetime.strptime(date_val.strip(), "%Y-%m-%d").date()
            else:
                errors.append(f"Row {r}: Invalid date format '{date_val}'")
                continue
        except (ValueError, AttributeError):
            errors.append(f"Row {r}: Could not parse date '{date_val}'")
            continue

        # Parse shift times
        start_str = str(ws[f"C{r}"].value or "").strip()
        end_str = str(ws[f"D{r}"].value or "").strip()
        if not start_str or not end_str:
            errors.append(f"Row {r}: Missing shift start or end time")
            continue

        try:
            start_time = datetime.strptime(start_str, "%H:%M")
            end_time = datetime.strptime(end_str, "%H:%M")
        except ValueError:
            errors.append(f"Row {r}: Invalid time format '{start_str}' or '{end_str}'. Use HH:MM.")
            continue

        shift_start = datetime.combine(shift_date, start_time.time())
        # Handle overnight shifts
        if end_time.time() <= start_time.time():
            shift_end = datetime.combine(shift_date + timedelta(days=1), end_time.time())
        else:
            shift_end = datetime.combine(shift_date, end_time.time())

        # Find employee
        guard_name = str(ws[f"E{r}"].value or "").strip()
        emp_id_val = ws[f"F{r}"].value
        employee = None

        if emp_id_val:
            try:
                employee = emp_by_id.get(int(emp_id_val))
            except (ValueError, TypeError):
                pass

        if not employee and guard_name:
            employee = emp_by_name.get(guard_name.lower())

        if not employee:
            if guard_name or emp_id_val:
                warnings.append(f"Row {r}: Employee not found (name='{guard_name}', id='{emp_id_val}'). Shift created without assignment.")
            else:
                # Empty row (just date/time template) — skip
                continue

        # Create or find shift
        shift_site_id = site.site_id if site else (site_id or None)
        if not shift_site_id:
            # Try to use first site for the org
            first_site = db.query(Site).filter(Site.org_id == org_id).first()
            if first_site:
                shift_site_id = first_site.site_id
            else:
                errors.append(f"Row {r}: No site found for this organization. Please provide site_id.")
                continue

        shift = Shift(
            org_id=org_id,
            site_id=shift_site_id,
            start_time=shift_start,
            end_time=shift_end,
            required_staff=1,
            status=ShiftStatus.CONFIRMED,
            notes=str(ws[f"I{r}"].value or ""),
        )
        db.add(shift)
        db.flush()  # Get shift_id
        shifts_created += 1

        if employee:
            # Create shift assignment
            assignment = ShiftAssignment(
                shift_id=shift.shift_id,
                employee_id=employee.employee_id,
                roster_id=roster.roster_id,
                status=AssignmentStatus.CONFIRMED.value,
                is_confirmed=True,
                confirmation_datetime=now,
                assigned_by=current_user.user_id,
            )

            # Calculate cost
            try:
                assignment.calculate_cost(shift, employee)
            except Exception as e:
                logger.warning(f"Row {r}: Cost calculation failed: {e}")
                # Set basic hours at least
                duration = (shift_end - shift_start).total_seconds() / 3600
                assignment.regular_hours = duration
                assignment.regular_pay = duration * (employee.hourly_rate or 0)
                assignment.total_cost = assignment.regular_pay

            total_cost += assignment.total_cost
            db.add(assignment)
            assignments_created += 1

            # PSIRA validation
            psira_num = str(ws[f"G{r}"].value or "").strip()
            if psira_num and employee.psira_number and psira_num != employee.psira_number:
                warnings.append(f"Row {r}: PSIRA number mismatch — template: {psira_num}, system: {employee.psira_number}")

    # Update roster stats
    roster.total_shifts = shifts_created
    roster.assigned_shifts = assignments_created
    roster.unassigned_shifts = shifts_created - assignments_created
    roster.total_cost = total_cost

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"Roster upload commit failed: {e}")
        raise HTTPException(status_code=500, detail="Failed to save roster. Please try again.")

    return {
        "status": "success",
        "roster_id": roster.roster_id,
        "roster_code": roster.roster_code,
        "period": f"{period_start} to {period_end}",
        "summary": {
            "rows_processed": rows_processed,
            "shifts_created": shifts_created,
            "assignments_created": assignments_created,
            "total_cost": round(total_cost, 2),
            "warnings_count": len(warnings),
            "errors_count": len(errors),
        },
        "warnings": warnings[:50],  # Limit to first 50
        "errors": errors[:50],
    }
