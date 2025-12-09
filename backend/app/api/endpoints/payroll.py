"""Payroll API endpoints with SA deductions integration."""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_
from datetime import date, datetime, timedelta
from typing import List, Optional
from pydantic import BaseModel
import io

from app.database import get_db
from app.models.payroll import PayrollSummary
from app.models.shift import Shift
from app.models.shift_assignment import ShiftAssignment
from app.models.employee import Employee
from app.auth.security import get_current_org_id
from app.services.payroll_generator_service import PayrollGeneratorService, PayrollPeriod

router = APIRouter()


# Pydantic schemas
class PayrollCreate(BaseModel):
    employee_id: int
    period_start: date
    period_end: date


class ComprehensivePayrollRequest(BaseModel):
    """Request for comprehensive payroll generation with SA deductions."""
    # Support both naming conventions - frontend uses period_start/period_end
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    period_start: Optional[date] = None  # Alias for start_date
    period_end: Optional[date] = None    # Alias for end_date
    employee_ids: Optional[List[int]] = None  # None = all employees
    client_ids: Optional[List[int]] = None  # Filter by client
    include_deductions: bool = True
    period_type: str = "monthly"  # weekly, bi_weekly, monthly

    @property
    def effective_start_date(self) -> date:
        """Get start date from either field name."""
        return self.start_date or self.period_start

    @property
    def effective_end_date(self) -> date:
        """Get end date from either field name."""
        return self.end_date or self.period_end


@router.get("/")
async def get_payroll(
    employee_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 100,
    org_id: int = Depends(get_current_org_id),
    db: Session = Depends(get_db)
):
    """Get payroll records with optional filters (filtered by organization via employee)."""
    query = db.query(PayrollSummary).join(Employee).filter(Employee.org_id == org_id)

    if employee_id:
        query = query.filter(PayrollSummary.employee_id == employee_id)

    payrolls = query.order_by(PayrollSummary.period_start.desc()).offset(skip).limit(limit).all()

    # Add employee name
    result = []
    for p in payrolls:
        employee = db.query(Employee).filter(Employee.employee_id == p.employee_id).first()
        result.append({
            "payroll_id": p.payroll_id,
            "employee_id": p.employee_id,
            "employee_name": f"{employee.first_name} {employee.last_name}" if employee else "Unknown",
            "period_start": p.period_start.isoformat(),
            "period_end": p.period_end.isoformat(),
            "total_hours": p.total_hours,
            "overtime_hours": p.overtime_hours,
            "gross_pay": p.gross_pay,
            "expenses_total": p.expenses_total,
            "net_pay": p.net_pay
        })

    return result


@router.get("/current-period")
async def get_current_period_payroll(
    org_id: int = Depends(get_current_org_id),
    db: Session = Depends(get_db)
):
    """Get payroll for current pay period (current month, filtered by organization via employee)."""
    today = date.today()
    period_start = today.replace(day=1)

    # Get last day of month
    if today.month == 12:
        period_end = today.replace(day=31)
    else:
        period_end = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))

    payrolls = db.query(PayrollSummary).join(Employee).filter(
        Employee.org_id == org_id,
        PayrollSummary.period_start >= period_start,
        PayrollSummary.period_end <= period_end
    ).all()

    # Calculate totals
    total_gross = sum(p.gross_pay for p in payrolls)
    total_expenses = sum(p.expenses_total for p in payrolls)
    total_net = sum(p.net_pay for p in payrolls)
    total_hours = sum(p.total_hours for p in payrolls)

    return {
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "employee_count": len(payrolls),
        "total_hours": total_hours,
        "total_gross_pay": total_gross,
        "total_expenses": total_expenses,
        "total_net_pay": total_net
    }


@router.post("/generate")
async def generate_payroll(
    payroll_data: PayrollCreate,
    org_id: int = Depends(get_current_org_id),
    db: Session = Depends(get_db)
):
    """Generate payroll for an employee and period (employee must belong to organization)."""
    # Get employee and verify it belongs to organization
    employee = db.query(Employee).filter(
        Employee.employee_id == payroll_data.employee_id,
        Employee.org_id == org_id
    ).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found in your organization")

    # Get all shift assignments for employee in period
    # Using ShiftAssignment instead of Shift.assigned_employee_id (which was removed in Phase 2)
    assignments = db.query(ShiftAssignment).join(Shift).filter(
        and_(
            ShiftAssignment.employee_id == payroll_data.employee_id,
            Shift.start_time >= datetime.combine(payroll_data.period_start, datetime.min.time()),
            Shift.end_time <= datetime.combine(payroll_data.period_end, datetime.max.time())
        )
    ).all()

    # Calculate totals from shift assignments
    # ShiftAssignment already has cost breakdown calculated
    total_regular_hours = sum(a.regular_hours for a in assignments)
    total_overtime_hours = sum(a.overtime_hours for a in assignments)
    total_hours = total_regular_hours + total_overtime_hours

    # Sum up all pay components
    regular_pay = sum(a.regular_pay for a in assignments)
    overtime_pay = sum(a.overtime_pay for a in assignments)
    night_premium = sum(a.night_premium for a in assignments)
    weekend_premium = sum(a.weekend_premium for a in assignments)
    travel_reimbursement = sum(a.travel_reimbursement for a in assignments)

    gross_pay = regular_pay + overtime_pay + night_premium + weekend_premium + travel_reimbursement

    # MVP: No additional expenses or deductions
    expenses_total = travel_reimbursement  # Travel is part of total cost
    net_pay = gross_pay

    # Create or update payroll record
    existing = db.query(PayrollSummary).filter(
        and_(
            PayrollSummary.employee_id == payroll_data.employee_id,
            PayrollSummary.period_start == payroll_data.period_start,
            PayrollSummary.period_end == payroll_data.period_end
        )
    ).first()

    if existing:
        existing.total_hours = total_hours
        existing.overtime_hours = total_overtime_hours
        existing.gross_pay = gross_pay
        existing.expenses_total = expenses_total
        existing.net_pay = net_pay
        db.commit()
        db.refresh(existing)
        payroll = existing
    else:
        payroll = PayrollSummary(
            employee_id=payroll_data.employee_id,
            period_start=payroll_data.period_start,
            period_end=payroll_data.period_end,
            total_hours=total_hours,
            overtime_hours=total_overtime_hours,
            gross_pay=gross_pay,
            expenses_total=expenses_total,
            net_pay=net_pay
        )
        db.add(payroll)
        db.commit()
        db.refresh(payroll)

    return {
        "payroll_id": payroll.payroll_id,
        "employee_id": payroll.employee_id,
        "period_start": payroll.period_start.isoformat(),
        "period_end": payroll.period_end.isoformat(),
        "total_hours": payroll.total_hours,
        "regular_hours": total_regular_hours,
        "overtime_hours": payroll.overtime_hours,
        "regular_pay": regular_pay,
        "overtime_pay": overtime_pay,
        "night_premium": night_premium,
        "weekend_premium": weekend_premium,
        "travel_reimbursement": travel_reimbursement,
        "gross_pay": payroll.gross_pay,
        "expenses_total": payroll.expenses_total,
        "net_pay": payroll.net_pay,
        "shift_count": len(assignments)
    }


@router.get("/{payroll_id}")
async def get_payroll_detail(
    payroll_id: int,
    org_id: int = Depends(get_current_org_id),
    db: Session = Depends(get_db)
):
    """Get detailed payroll record (filtered by organization via employee)."""
    payroll = db.query(PayrollSummary).join(Employee).filter(
        PayrollSummary.payroll_id == payroll_id,
        Employee.org_id == org_id
    ).first()

    if not payroll:
        raise HTTPException(status_code=404, detail="Payroll record not found")

    employee = db.query(Employee).filter(Employee.employee_id == payroll.employee_id).first()

    return {
        "payroll_id": payroll.payroll_id,
        "employee": {
            "employee_id": employee.employee_id,
            "name": f"{employee.first_name} {employee.last_name}",
            "hourly_rate": employee.hourly_rate
        } if employee else None,
        "period_start": payroll.period_start.isoformat(),
        "period_end": payroll.period_end.isoformat(),
        "total_hours": payroll.total_hours,
        "overtime_hours": payroll.overtime_hours,
        "gross_pay": payroll.gross_pay,
        "expenses_total": payroll.expenses_total,
        "net_pay": payroll.net_pay
    }


@router.delete("/{payroll_id}")
async def delete_payroll(
    payroll_id: int,
    org_id: int = Depends(get_current_org_id),
    db: Session = Depends(get_db)
):
    """Delete payroll record (filtered by organization via employee)."""
    payroll = db.query(PayrollSummary).join(Employee).filter(
        PayrollSummary.payroll_id == payroll_id,
        Employee.org_id == org_id
    ).first()

    if not payroll:
        raise HTTPException(status_code=404, detail="Payroll record not found")

    db.delete(payroll)
    db.commit()

    return {"message": "Payroll deleted successfully"}


# =============================================================================
# COMPREHENSIVE PAYROLL WITH SA DEDUCTIONS
# =============================================================================

@router.post("/generate-comprehensive")
async def generate_comprehensive_payroll(
    request: ComprehensivePayrollRequest,
    org_id: int = Depends(get_current_org_id),
    db: Session = Depends(get_db)
):
    """
    Generate comprehensive payroll for all/selected employees with SA deductions.

    This endpoint integrates:
    - Hours calculation from shift assignments
    - Night shift premiums (10%)
    - Sunday premiums (50% extra)
    - Holiday premiums (100% extra)
    - Overtime calculation
    - PAYE tax (per SARS 2024/2025 brackets)
    - UIF (1% employee, 1% employer, capped at R177.12)
    - SDL (1% employer)
    - Other deductions (PSIRA, Bargaining Council, Provident Fund)

    Returns complete payroll data matching the client Excel format.
    """
    # Validate dates
    start = request.effective_start_date
    end = request.effective_end_date
    if not start or not end:
        raise HTTPException(status_code=400, detail="Start and end dates are required")

    try:
        payroll_data = PayrollGeneratorService.generate_payroll(
            db=db,
            org_id=org_id,
            start_date=start,
            end_date=end,
            employee_ids=request.employee_ids,
            client_ids=request.client_ids,
            include_deductions=request.include_deductions,
            period_type=request.period_type
        )

        if payroll_data.get("status") == "error":
            raise HTTPException(status_code=400, detail=payroll_data.get("message"))

        return payroll_data

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating payroll: {str(e)}")


@router.post("/generate-comprehensive/excel")
async def generate_payroll_excel(
    request: ComprehensivePayrollRequest,
    org_id: int = Depends(get_current_org_id),
    db: Session = Depends(get_db)
):
    """
    Generate comprehensive payroll and export as Excel file.

    Format matches the ATHOPASI SECURITY SERVICES payroll template:
    - Employee details (ID, Name, Bank, Tax)
    - Hours breakdown (Normal, Sunday, Holiday, Extra)
    - Earnings (Wage, Allowances, Bonus)
    - Deductions (UIF, PAYE, PSIRA, Provident)
    - Net Salary
    """
    # Validate dates
    start = request.effective_start_date
    end = request.effective_end_date
    if not start or not end:
        raise HTTPException(status_code=400, detail="Start and end dates are required")

    try:
        # Import openpyxl for Excel generation
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="openpyxl not installed. Run: pip install openpyxl"
            )

        # Generate payroll data
        payroll_data = PayrollGeneratorService.generate_payroll(
            db=db,
            org_id=org_id,
            start_date=start,
            end_date=end,
            employee_ids=request.employee_ids,
            client_ids=request.client_ids,
            include_deductions=request.include_deductions,
            period_type=request.period_type
        )

        if payroll_data.get("status") == "error":
            raise HTTPException(status_code=400, detail=payroll_data.get("message"))

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Payroll {request.start_date.strftime('%Y-%m')}"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
        money_format = '#,##0.00'
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Define columns matching client Excel format
        columns = [
            "NO.", "SURNAME", "NAME(S)", "ID NUMBER", "GENDER", "TAX NUMBER", "PSIRA",
            "BANK NAME(S)", "ACCOUNT NUMBERS", "EMPLOYEE NO.", "TOTAL DAYS W",
            "NORMAL HRS", "SUNDAY HRS", "HOLIDAY HRS", "EXTRA HRS", "TOTAL HRS W",
            "RATE P/HR", "TOTAL HRS WAGE", "NIGHT S/ALL", "SUPERVISOR", "FUNERAL",
            "GUN", "BONUS", "CLEANING", "OVERTIME", "SICK", "GROSS SALARY",
            "UIF", "PSIRA", "BARGAINING", "PAYE", "PROFIDENT FUND", "NUCAAW",
            "HOSP COVER", "DEFFECT", "NET SALARY"
        ]

        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        # Write employee data
        for row_idx, emp in enumerate(payroll_data["employees"], 2):
            row_data = [
                row_idx - 1,  # NO.
                emp["surname"],
                emp["names"],
                emp["id_number"],
                emp["gender"],
                emp["tax_number"],
                emp["psira_number"],
                emp["bank_name"],
                emp["account_number_full"],  # Full account for internal use
                emp["employee_no"],
                emp["total_days_worked"],
                emp["normal_hours"],
                emp["sunday_hours"],
                emp["holiday_hours"],
                emp["extra_hours"],
                emp["total_hours_worked"],
                emp["rate_per_hour"],
                emp["total_hours_wage"],
                emp["night_allowance"],
                emp["supervisor_allowance"],
                emp["funeral_allowance"],
                emp["gun_allowance"],
                emp["bonus"],
                emp["cleaning_allowance"],
                emp["overtime_pay"],
                emp["sick_pay"],
                emp["gross_salary"],
                emp["uif"],
                emp["psira_deduction"],
                emp["bargaining_council"],
                emp["paye"],
                emp["provident_fund"],
                emp["nucaaw"],
                emp["hospital_cover"],
                emp["defect"],
                emp["net_salary"]
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Format money columns
                if col_idx >= 17 and col_idx <= 36:  # Rate and all money columns
                    cell.number_format = money_format
                    cell.alignment = Alignment(horizontal='right')

        # Add summary row
        summary_row = len(payroll_data["employees"]) + 3
        ws.cell(row=summary_row, column=1, value="TOTALS:").font = Font(bold=True)
        ws.cell(row=summary_row, column=27, value=payroll_data["summary"]["total_gross_pay"]).font = Font(bold=True)
        ws.cell(row=summary_row, column=36, value=payroll_data["summary"]["total_net_pay"]).font = Font(bold=True)

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 2, 12)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Payroll_{request.start_date.strftime('%Y-%m-%d')}_to_{request.end_date.strftime('%Y-%m-%d')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")


@router.get("/sa-tax-tables")
async def get_sa_tax_tables():
    """
    Get South African tax tables reference for 2024/2025 tax year.

    Useful for displaying tax information to users.
    """
    from app.services.sa_payroll_service import get_tax_tables_reference
    return get_tax_tables_reference()


@router.post("/calculate-employee-payslip")
async def calculate_employee_payslip(
    employee_id: int,
    start_date: date,
    end_date: date,
    org_id: int = Depends(get_current_org_id),
    db: Session = Depends(get_db)
):
    """
    Calculate detailed payslip for a single employee.

    Returns complete breakdown with:
    - Hours worked (regular, overtime, night, weekend, holiday)
    - Earnings (gross pay, premiums, allowances)
    - Deductions (PAYE, UIF, other)
    - Net pay
    - Employer contributions
    """
    payroll_data = PayrollGeneratorService.generate_payroll(
        db=db,
        org_id=org_id,
        start_date=start_date,
        end_date=end_date,
        employee_ids=[employee_id],
        include_deductions=True
    )

    if payroll_data.get("status") == "error":
        raise HTTPException(status_code=400, detail=payroll_data.get("message"))

    if not payroll_data["employees"]:
        raise HTTPException(status_code=404, detail="Employee not found or no shifts in period")

    return {
        "payroll_id": payroll_data["payroll_id"],
        "period": payroll_data["period"],
        "employee": payroll_data["employees"][0],
        "employer_contributions": {
            "uif": payroll_data["summary"]["employer_contributions"]["uif"],
            "sdl": payroll_data["summary"]["employer_contributions"]["sdl"]
        }
    }


# =============================================================================
# PAYSLIP PDF GENERATION
# =============================================================================

class PayslipPDFRequest(BaseModel):
    """Request for generating payslip PDF."""
    employee_id: int
    start_date: date
    end_date: date
    payment_date: Optional[date] = None  # Defaults to end_date + 7 days


class BulkPayslipPDFRequest(BaseModel):
    """Request for generating bulk payslip PDFs."""
    # Support both date-based and ID-based bulk generation
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    payment_date: Optional[date] = None
    employee_ids: Optional[List[int]] = None  # None = all employees
    payroll_ids: Optional[List[int]] = None  # Alternative: by payroll record IDs


@router.post("/payslip/pdf")
async def generate_payslip_pdf(
    request: PayslipPDFRequest,
    org_id: int = Depends(get_current_org_id),
    db: Session = Depends(get_db)
):
    """
    Generate PDF payslip for a single employee.

    Returns a professional PDF payslip with:
    - Company header and details
    - Employee information (masked banking details)
    - Earnings breakdown (basic, overtime, premiums, allowances)
    - Deductions breakdown (PAYE, UIF, PSIRA, etc.)
    - Net pay prominently displayed
    - Optional YTD summary and leave balances
    """
    from app.services.payslip_generator import (
        PayslipPDFGenerator, create_payslip_from_payroll_record
    )
    from app.models.organization import Organization

    # Get organization details
    org = db.query(Organization).filter(Organization.org_id == org_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    # Generate payroll data for employee
    payroll_data = PayrollGeneratorService.generate_payroll(
        db=db,
        org_id=org_id,
        start_date=request.start_date,
        end_date=request.end_date,
        employee_ids=[request.employee_id],
        include_deductions=True
    )

    if payroll_data.get("status") == "error":
        raise HTTPException(status_code=400, detail=payroll_data.get("message"))

    if not payroll_data["employees"]:
        raise HTTPException(status_code=404, detail="Employee not found or no shifts in period")

    emp_data = payroll_data["employees"][0]

    # Company information
    company_info = {
        "company_name": org.company_name,
        "registration_number": "",  # Add if available in org model
        "psira_registration": org.psira_company_registration or "",
        "vat_number": "",  # Add if available
        "address_line1": "",  # Add if available
        "city": "",
        "postal_code": "",
        "phone": "",
        "email": org.billing_email or ""
    }

    # Payment date (default to 7 days after period end)
    payment_date = request.payment_date or (request.end_date + timedelta(days=7))

    # Create payslip data
    payslip_data = create_payslip_from_payroll_record(
        employee_record=emp_data,
        company_info=company_info,
        period_start=request.start_date,
        period_end=request.end_date,
        payment_date=payment_date,
        payslip_number=f"PS-{org_id}-{request.employee_id}-{request.end_date.strftime('%Y%m')}"
    )

    # Generate PDF
    generator = PayslipPDFGenerator()
    pdf_buffer = generator.generate_pdf(payslip_data)

    filename = f"Payslip_{emp_data['surname']}_{emp_data['names']}_{request.start_date.strftime('%Y%m%d')}_{request.end_date.strftime('%Y%m%d')}.pdf"

    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/payslips/pdf/bulk")
async def generate_bulk_payslips_pdf(
    request: BulkPayslipPDFRequest,
    org_id: int = Depends(get_current_org_id),
    db: Session = Depends(get_db)
):
    """
    Generate PDF payslips for multiple employees in a single document.

    Supports two modes:
    1. By payroll_ids: Generate PDFs from existing payroll records
    2. By date range: Generate PDFs by calculating payroll for date range

    Each employee gets a separate page in the PDF.
    Useful for batch printing or distribution.
    """
    from app.services.payslip_generator import (
        PayslipPDFGenerator, create_payslip_from_payroll_record
    )
    from app.models.organization import Organization

    # Get organization details
    org = db.query(Organization).filter(Organization.org_id == org_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    # Company information
    company_info = {
        "company_name": org.company_name,
        "registration_number": "",
        "psira_registration": org.psira_company_registration or "",
        "vat_number": "",
        "address_line1": "",
        "city": "",
        "postal_code": "",
        "phone": "",
        "email": org.billing_email or ""
    }

    payslips = []

    # Mode 1: Generate by payroll_ids (from existing payroll records)
    if request.payroll_ids:
        for payroll_id in request.payroll_ids:
            # Get payroll record
            payroll = db.query(PayrollSummary).join(Employee).filter(
                PayrollSummary.payroll_id == payroll_id,
                Employee.org_id == org_id
            ).first()

            if not payroll:
                continue  # Skip records not found

            employee = db.query(Employee).filter(
                Employee.employee_id == payroll.employee_id
            ).first()

            if not employee:
                continue

            # Generate payroll data
            payroll_data = PayrollGeneratorService.generate_payroll(
                db=db,
                org_id=org_id,
                start_date=payroll.period_start,
                end_date=payroll.period_end,
                employee_ids=[employee.employee_id],
                include_deductions=True
            )

            if payroll_data.get("employees"):
                emp_data = payroll_data["employees"][0]
            else:
                # Fall back to basic data
                emp_data = {
                    "employee_id": employee.employee_id,
                    "surname": employee.last_name,
                    "names": employee.first_name,
                    "id_number": employee.id_number or "",
                    "gender": employee.gender or "",
                    "tax_number": employee.tax_number or "",
                    "psira_number": "",
                    "bank_name": employee.bank_name or "",
                    "account_number_full": employee.account_number or "",
                    "employee_no": employee.employee_number or f"EMP{employee.employee_id:04d}",
                    "total_days_worked": int(payroll.total_hours / 8) if payroll.total_hours else 0,
                    "normal_hours": (payroll.total_hours or 0) - (payroll.overtime_hours or 0),
                    "sunday_hours": 0,
                    "holiday_hours": 0,
                    "extra_hours": payroll.overtime_hours or 0,
                    "total_hours_worked": payroll.total_hours or 0,
                    "rate_per_hour": employee.hourly_rate or 0,
                    "total_hours_wage": payroll.gross_pay or 0,
                    "night_allowance": 0,
                    "supervisor_allowance": 0,
                    "funeral_allowance": 0,
                    "gun_allowance": 0,
                    "bonus": 0,
                    "cleaning_allowance": 0,
                    "overtime_pay": 0,
                    "sick_pay": 0,
                    "gross_salary": payroll.gross_pay or 0,
                    "uif": 0,
                    "psira_deduction": 0,
                    "bargaining_council": 0,
                    "paye": 0,
                    "provident_fund": 0,
                    "nucaaw": 0,
                    "hospital_cover": 0,
                    "defect": 0,
                    "net_salary": payroll.net_pay or 0
                }

            payment_date = request.payment_date or (payroll.period_end + timedelta(days=7))

            payslip_data = create_payslip_from_payroll_record(
                employee_record=emp_data,
                company_info=company_info,
                period_start=payroll.period_start,
                period_end=payroll.period_end,
                payment_date=payment_date,
                payslip_number=f"PS-{org_id}-{employee.employee_id}-{payroll.period_end.strftime('%Y%m')}"
            )
            payslips.append(payslip_data)

        if not payslips:
            raise HTTPException(status_code=404, detail="No valid payroll records found")

        filename = f"Payslips_Bulk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    # Mode 2: Generate by date range
    elif request.start_date and request.end_date:
        payroll_data = PayrollGeneratorService.generate_payroll(
            db=db,
            org_id=org_id,
            start_date=request.start_date,
            end_date=request.end_date,
            employee_ids=request.employee_ids,
            include_deductions=True
        )

        if payroll_data.get("status") == "error":
            raise HTTPException(status_code=400, detail=payroll_data.get("message"))

        if not payroll_data["employees"]:
            raise HTTPException(status_code=404, detail="No employees found or no shifts in period")

        payment_date = request.payment_date or (request.end_date + timedelta(days=7))

        for emp_data in payroll_data["employees"]:
            payslip_data = create_payslip_from_payroll_record(
                employee_record=emp_data,
                company_info=company_info,
                period_start=request.start_date,
                period_end=request.end_date,
                payment_date=payment_date,
                payslip_number=f"PS-{org_id}-{emp_data['employee_id']}-{request.end_date.strftime('%Y%m')}"
            )
            payslips.append(payslip_data)

        filename = f"Payslips_Bulk_{request.start_date.strftime('%Y%m%d')}_{request.end_date.strftime('%Y%m%d')}.pdf"

    else:
        raise HTTPException(
            status_code=400,
            detail="Either payroll_ids or start_date/end_date must be provided"
        )

    # Generate bulk PDF
    generator = PayslipPDFGenerator()
    pdf_buffer = generator.generate_bulk_payslips(payslips)

    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/payslip/preview/{employee_id}")
async def preview_payslip(
    employee_id: int,
    start_date: date,
    end_date: date,
    org_id: int = Depends(get_current_org_id),
    db: Session = Depends(get_db)
):
    """
    Get payslip data in JSON format for UI preview.

    Returns formatted payslip data that can be displayed in the frontend
    before generating the PDF.
    """
    from app.models.organization import Organization

    # Get organization details
    org = db.query(Organization).filter(Organization.org_id == org_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    # Generate payroll data
    payroll_data = PayrollGeneratorService.generate_payroll(
        db=db,
        org_id=org_id,
        start_date=start_date,
        end_date=end_date,
        employee_ids=[employee_id],
        include_deductions=True
    )

    if payroll_data.get("status") == "error":
        raise HTTPException(status_code=400, detail=payroll_data.get("message"))

    if not payroll_data["employees"]:
        raise HTTPException(status_code=404, detail="Employee not found or no shifts in period")

    emp = payroll_data["employees"][0]

    # Return formatted preview data
    return {
        "company": {
            "name": org.company_name,
            "psira_registration": org.psira_company_registration,
            "billing_email": org.billing_email
        },
        "period": {
            "start": start_date.isoformat(),
            "end": end_date.isoformat()
        },
        "employee": {
            "id": emp["employee_id"],
            "employee_number": emp["employee_no"],
            "name": f"{emp['names']} {emp['surname']}",
            "id_number_masked": emp["id_number"][:6] + "****" + emp["id_number"][-3:] if emp["id_number"] else "",
            "psira_number": emp["psira_number"],
            "bank_name": emp["bank_name"],
            "account_number_masked": "****" + emp["account_number_full"][-4:] if emp["account_number_full"] else "",
            "tax_number": emp["tax_number"]
        },
        "hours": {
            "normal": emp["normal_hours"],
            "overtime": emp["extra_hours"],
            "sunday": emp["sunday_hours"],
            "holiday": emp["holiday_hours"],
            "total": emp["total_hours_worked"]
        },
        "earnings": {
            "basic_salary": emp["total_hours_wage"],
            "overtime_pay": emp["overtime_pay"],
            "night_allowance": emp["night_allowance"],
            "sunday_premium": emp.get("sunday_premium", 0),
            "holiday_premium": emp.get("holiday_premium", 0),
            "supervisor_allowance": emp["supervisor_allowance"],
            "travel_allowance": emp.get("travel_allowance", 0),
            "gun_allowance": emp["gun_allowance"],
            "cleaning_allowance": emp["cleaning_allowance"],
            "bonus": emp["bonus"],
            "sick_pay": emp["sick_pay"],
            "gross_salary": emp["gross_salary"]
        },
        "deductions": {
            "paye": emp["paye"],
            "uif": emp["uif"],
            "psira": emp["psira_deduction"],
            "bargaining_council": emp["bargaining_council"],
            "provident_fund": emp["provident_fund"],
            "hospital_cover": emp["hospital_cover"],
            "nucaaw": emp["nucaaw"],
            "defect": emp["defect"],
            "total": emp["paye"] + emp["uif"] + emp["psira_deduction"] + emp["bargaining_council"] + emp["provident_fund"]
        },
        "net_salary": emp["net_salary"],
        "currency": "ZAR"
    }


# =============================================================================
# PAYSLIP PDF BY PAYROLL ID (Frontend expects this endpoint)
# =============================================================================

@router.get("/{payroll_id}/payslip/pdf")
async def get_payslip_pdf_by_payroll_id(
    payroll_id: int,
    org_id: int = Depends(get_current_org_id),
    db: Session = Depends(get_db)
):
    """
    Generate PDF payslip for an existing payroll record.

    This endpoint is called by the frontend after payroll has been generated.
    It retrieves the payroll record and generates a PDF from it.
    """
    from app.services.payslip_generator import (
        PayslipPDFGenerator, create_payslip_from_payroll_record
    )
    from app.models.organization import Organization

    # Get payroll record with organization filter
    payroll = db.query(PayrollSummary).join(Employee).filter(
        PayrollSummary.payroll_id == payroll_id,
        Employee.org_id == org_id
    ).first()

    if not payroll:
        raise HTTPException(status_code=404, detail="Payroll record not found")

    # Get employee details
    employee = db.query(Employee).filter(
        Employee.employee_id == payroll.employee_id
    ).first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Get organization details
    org = db.query(Organization).filter(Organization.org_id == org_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    # Regenerate payroll data for this employee and period
    payroll_data = PayrollGeneratorService.generate_payroll(
        db=db,
        org_id=org_id,
        start_date=payroll.period_start,
        end_date=payroll.period_end,
        employee_ids=[employee.employee_id],
        include_deductions=True
    )

    if payroll_data.get("status") == "error" or not payroll_data.get("employees"):
        # Fall back to basic payroll record data
        emp_data = {
            "employee_id": employee.employee_id,
            "surname": employee.last_name,
            "names": employee.first_name,
            "id_number": employee.id_number or "",
            "gender": employee.gender or "",
            "tax_number": employee.tax_number or "",
            "psira_number": "",
            "bank_name": employee.bank_name or "",
            "account_number_full": employee.account_number or "",
            "employee_no": employee.employee_number or f"EMP{employee.employee_id:04d}",
            "total_days_worked": int(payroll.total_hours / 8),
            "normal_hours": payroll.total_hours - payroll.overtime_hours,
            "sunday_hours": 0,
            "holiday_hours": 0,
            "extra_hours": payroll.overtime_hours,
            "total_hours_worked": payroll.total_hours,
            "rate_per_hour": employee.hourly_rate or 0,
            "total_hours_wage": payroll.gross_pay,
            "night_allowance": 0,
            "supervisor_allowance": 0,
            "funeral_allowance": 0,
            "gun_allowance": 0,
            "bonus": 0,
            "cleaning_allowance": 0,
            "overtime_pay": 0,
            "sick_pay": 0,
            "gross_salary": payroll.gross_pay,
            "uif": 0,
            "psira_deduction": 0,
            "bargaining_council": 0,
            "paye": 0,
            "provident_fund": 0,
            "nucaaw": 0,
            "hospital_cover": 0,
            "defect": 0,
            "net_salary": payroll.net_pay
        }
    else:
        emp_data = payroll_data["employees"][0]

    # Company information
    company_info = {
        "company_name": org.company_name,
        "registration_number": "",
        "psira_registration": org.psira_company_registration or "",
        "vat_number": "",
        "address_line1": "",
        "city": "",
        "postal_code": "",
        "phone": "",
        "email": org.billing_email or ""
    }

    # Payment date (7 days after period end)
    payment_date = payroll.period_end + timedelta(days=7)

    # Create payslip data
    payslip_data = create_payslip_from_payroll_record(
        employee_record=emp_data,
        company_info=company_info,
        period_start=payroll.period_start,
        period_end=payroll.period_end,
        payment_date=payment_date,
        payslip_number=f"PS-{org_id}-{employee.employee_id}-{payroll.period_end.strftime('%Y%m')}"
    )

    # Generate PDF
    generator = PayslipPDFGenerator()
    pdf_buffer = generator.generate_pdf(payslip_data)

    filename = f"Payslip_{emp_data['surname']}_{emp_data['names']}_{payroll.period_start.strftime('%Y%m%d')}_{payroll.period_end.strftime('%Y%m%d')}.pdf"

    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
