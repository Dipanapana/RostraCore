"""
Mafoko Security Site Roster Import Service.

Parses Mafoko's D/o grid Excel format and creates Roster + Shift + ShiftAssignment records.

Mafoko format:
- Header: "Mafoko Security Services - {Province} Site Roster for {Client} {SiteCode}: {SiteName} {Schedule}"
- Columns: Pers No (MSP#####), Grade, Name (SURNAME,INITIAL), Contact, Worked, then 31 day columns (1-31)
- Cell values: "D" = day shift, "o" = off
"""

import hashlib
import io
import logging
import re
import uuid
from datetime import date, datetime, time, timedelta
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.client import Client
from app.models.employee import Employee, EmployeeStatus
from app.models.roster import Roster
from app.models.roster_upload_log import RosterUploadLog
from app.models.shift import Shift, ShiftStatus
from app.models.shift_assignment import ShiftAssignment, AssignmentStatus
from app.models.site import Site

logger = logging.getLogger(__name__)


# Regex to parse Mafoko header line
HEADER_REGEX = re.compile(
    r"Mafoko\s+Security\s+(?:Services|Patrols)\s*[-–—]\s*"
    r"(?P<province>[\w\s]+?)\s+"
    r"Site\s+Roster\s+for\s+"
    r"(?P<client>[\w\s&]+?)\s+"
    r"(?P<site_code>\d+)\s*:\s*"
    r"(?P<site_name>[\w\s]+?)\s+"
    r"(?P<schedule>[\w\-/]+(?:\s+[\w\-/]+)?)",
    re.IGNORECASE,
)

# Regex for date range from filename: "21 MAY TO 20 JUNE 2025"
FILENAME_DATE_REGEX = re.compile(
    r"(\d{1,2})\s+"
    r"(JAN(?:UARY)?|FEB(?:RUARY)?|MAR(?:CH)?|APR(?:IL)?|MAY|JUN(?:E)?|JUL(?:Y)?|AUG(?:UST)?|SEP(?:TEMBER)?|OCT(?:OBER)?|NOV(?:EMBER)?|DEC(?:EMBER)?)\s+"
    r"TO\s+"
    r"(\d{1,2})\s+"
    r"(JAN(?:UARY)?|FEB(?:RUARY)?|MAR(?:CH)?|APR(?:IL)?|MAY|JUN(?:E)?|JUL(?:Y)?|AUG(?:UST)?|SEP(?:TEMBER)?|OCT(?:OBER)?|NOV(?:EMBER)?|DEC(?:EMBER)?)\s+"
    r"(\d{4})",
    re.IGNORECASE,
)

MONTH_MAP = {
    "JAN": 1, "JANUARY": 1, "FEB": 2, "FEBRUARY": 2,
    "MAR": 3, "MARCH": 3, "APR": 4, "APRIL": 4,
    "MAY": 5, "JUN": 6, "JUNE": 6, "JUL": 7, "JULY": 7,
    "AUG": 8, "AUGUST": 8, "SEP": 9, "SEPTEMBER": 9,
    "OCT": 10, "OCTOBER": 10, "NOV": 11, "NOVEMBER": 11,
    "DEC": 12, "DECEMBER": 12,
}


class MafokoRosterImportService:
    """Parses Mafoko site roster Excel files and creates roster records."""

    @staticmethod
    def compute_file_hash(file_content: bytes) -> str:
        """Compute SHA-256 hash of file content for duplicate detection."""
        return hashlib.sha256(file_content).hexdigest()

    @staticmethod
    def detect_mafoko_format(ws) -> bool:
        """Check if a worksheet is in Mafoko Site Roster format."""
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(row=r, column=c).value
                if val and "mafoko" in str(val).lower() and "roster" in str(val).lower():
                    return True
        # Also check for "Pers No" column header
        for r in range(1, min(15, ws.max_row + 1)):
            val = ws.cell(row=r, column=1).value
            if val and str(val).strip().lower() in ("pers no", "pers no.", "personnel no"):
                return True
        return False

    @staticmethod
    def parse_header(ws) -> dict:
        """
        Parse the header row to extract province, client, site_code, site_name, schedule.
        Scans first 10 rows for the Mafoko header pattern.
        """
        header_info = {
            "province": None,
            "client": None,
            "site_code": None,
            "site_name": None,
            "schedule": None,
            "raw_header": None,
        }

        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(row=r, column=c).value
                if not val:
                    continue
                val_str = str(val).strip()
                if "mafoko" not in val_str.lower():
                    continue

                header_info["raw_header"] = val_str
                match = HEADER_REGEX.search(val_str)
                if match:
                    header_info["province"] = match.group("province").strip()
                    header_info["client"] = match.group("client").strip()
                    header_info["site_code"] = match.group("site_code").strip()
                    header_info["site_name"] = match.group("site_name").strip()
                    header_info["schedule"] = match.group("schedule").strip()
                return header_info

        return header_info

    # Regex for ISO date range in spreadsheet: "Scheduled Period: 2025-05-21 to 2025-06-20"
    SCHEDULED_PERIOD_REGEX = re.compile(
        r"(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})",
        re.IGNORECASE,
    )

    @classmethod
    def parse_date_range_from_sheet(cls, ws) -> tuple[Optional[date], Optional[date]]:
        """Try to parse date range from a 'Scheduled Period:' row in the worksheet."""
        for r in range(1, min(10, ws.max_row + 1)):
            val = ws.cell(row=r, column=1).value
            if not val:
                continue
            val_str = str(val).strip()
            match = cls.SCHEDULED_PERIOD_REGEX.search(val_str)
            if match:
                try:
                    start = date.fromisoformat(match.group(1))
                    end = date.fromisoformat(match.group(2))
                    return start, end
                except ValueError:
                    pass
        return None, None

    @staticmethod
    def parse_date_range(filename: str) -> tuple[Optional[date], Optional[date]]:
        """Extract start_date, end_date from filename like '21 MAY TO 20 JUNE 2025'."""
        match = FILENAME_DATE_REGEX.search(filename.upper())
        if not match:
            return None, None

        start_day = int(match.group(1))
        start_month = MONTH_MAP.get(match.group(2).upper())
        end_day = int(match.group(3))
        end_month = MONTH_MAP.get(match.group(4).upper())
        year = int(match.group(5))

        if not start_month or not end_month:
            return None, None

        # Handle cross-year: if start month is after end month, start is previous year
        start_year = year if start_month <= end_month else year - 1

        try:
            start_date = date(start_year, start_month, start_day)
            end_date = date(year, end_month, end_day)
            return start_date, end_date
        except ValueError:
            return None, None

    # Regex to match day column headers like "We 21", "Th 22", "Mo 01", etc.
    DAY_COL_REGEX = re.compile(
        r"^(?:Mo|Tu|We|Th|Fr|Sa|Su|Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+(\d{1,2})$",
        re.IGNORECASE,
    )

    @staticmethod
    def find_data_start(ws) -> tuple[Optional[int], dict]:
        """
        Find the row where employee data starts and map column indices.
        Returns (header_row, column_map) where column_map maps names to 1-based column indices.
        """
        column_map = {}
        for r in range(1, min(20, ws.max_row + 1)):
            val = ws.cell(row=r, column=1).value
            if not val:
                continue
            val_str = str(val).strip().lower()
            if val_str in ("pers no", "pers no.", "personnel no", "no"):
                # Map known columns
                column_map["pers_no"] = 1
                for c in range(2, min(ws.max_column + 1, 50)):
                    col_val = ws.cell(row=r, column=c).value
                    if not col_val:
                        continue
                    col_str = str(col_val).strip().lower()
                    if col_str == "grade":
                        column_map["grade"] = c
                    elif col_str in ("name", "names", "name(s)"):
                        column_map["name"] = c
                    elif "contact" in col_str or col_str == "cell":
                        column_map["contact"] = c
                    elif col_str in ("worked", "days worked", "total"):
                        column_map["worked"] = c
                    else:
                        # Try to detect day columns — either pure numbers (1-31)
                        # or day-abbreviation format ("We 21", "Th 22", "Mo 01")
                        is_day_col = False
                        try:
                            day_num = int(col_str)
                            if 1 <= day_num <= 31:
                                is_day_col = True
                        except (ValueError, TypeError):
                            pass

                        if not is_day_col:
                            day_match = MafokoRosterImportService.DAY_COL_REGEX.match(
                                str(col_val).strip()
                            )
                            if day_match:
                                is_day_col = True

                        if is_day_col:
                            if "first_day_col" not in column_map:
                                column_map["first_day_col"] = c
                            column_map["last_day_col"] = c
                return r, column_map

        return None, {}

    def parse_roster_data(self, ws, header_row: int, column_map: dict) -> list[dict]:
        """
        Parse all employee rows from the worksheet.
        Returns list of dicts with pers_no, grade, name, contact, worked, working_days (list of column indices).
        """
        rows = []
        first_day = column_map.get("first_day_col")
        last_day = column_map.get("last_day_col")

        if not first_day or not last_day:
            return rows

        for r in range(header_row + 1, ws.max_row + 1):
            pers_no_val = ws.cell(row=r, column=column_map.get("pers_no", 1)).value
            if not pers_no_val:
                continue

            pers_no = str(pers_no_val).strip()
            # Skip empty, summary, or repeated header rows
            if not pers_no or pers_no.lower() in (
                "total", "totals", "grand total", "",
                "pers no", "pers no.", "personnel no", "no",
            ):
                continue

            name_col = column_map.get("name", 3)
            grade_col = column_map.get("grade", 2)
            contact_col = column_map.get("contact", 4)
            worked_col = column_map.get("worked", 5)

            name = str(ws.cell(row=r, column=name_col).value or "").strip()
            grade = str(ws.cell(row=r, column=grade_col).value or "").strip()
            contact = str(ws.cell(row=r, column=contact_col).value or "").strip()
            worked_val = ws.cell(row=r, column=worked_col).value

            # Parse D/o grid — collect which day columns have "D"
            working_days = []
            for c in range(first_day, last_day + 1):
                cell_val = ws.cell(row=r, column=c).value
                if cell_val and str(cell_val).strip().upper() == "D":
                    # Column offset from first_day_col gives the day number (0-indexed)
                    day_offset = c - first_day
                    working_days.append(day_offset)

            # Count worked days
            try:
                worked_count = int(worked_val) if worked_val else len(working_days)
            except (ValueError, TypeError):
                worked_count = len(working_days)

            rows.append({
                "pers_no": pers_no,
                "name": name,
                "grade": grade,
                "contact": contact,
                "worked_count": worked_count,
                "working_days": working_days,  # 0-indexed offsets from start_date
                "row_number": r,
            })

        return rows

    def match_employees(
        self, db: Session, org_id: int, roster_rows: list[dict]
    ) -> tuple[dict, list[dict]]:
        """
        Match MSP##### personnel IDs to Employee records.
        Returns (matched_map: {pers_no: Employee}, unmatched_list: [{pers_no, name, grade}]).
        """
        employees = db.query(Employee).filter(
            Employee.org_id == org_id,
            Employee.status == EmployeeStatus.ACTIVE,
        ).all()

        # Build lookup maps
        by_emp_number = {}
        for e in employees:
            if e.employee_number:
                by_emp_number[e.employee_number.strip().upper()] = e

        by_name = {}
        for e in employees:
            if e.last_name:
                # Mafoko format: "SURNAME,INITIAL" or "SURNAME, INITIAL"
                key = e.last_name.upper().strip()
                if e.first_name:
                    key += f",{e.first_name[0].upper()}"
                by_name[key] = e
                # Also store without comma spacing variations
                by_name[key.replace(" ", "")] = e

        matched = {}
        unmatched = []

        for row in roster_rows:
            pers_no = row["pers_no"].strip().upper()
            name = row["name"]
            employee = None

            # Primary: match by employee_number
            employee = by_emp_number.get(pers_no)

            # Secondary: match by SURNAME,INITIAL
            if not employee and name:
                name_key = name.upper().strip().replace(" ", "")
                employee = by_name.get(name_key)

            if employee:
                matched[row["pers_no"]] = employee
            else:
                unmatched.append({
                    "pers_no": row["pers_no"],
                    "name": row["name"],
                    "grade": row["grade"],
                    "row_number": row["row_number"],
                })

        return matched, unmatched

    def resolve_site(
        self,
        db: Session,
        org_id: int,
        parsed_header: dict,
        site_id_override: Optional[int] = None,
        client_id_override: Optional[int] = None,
        auto_create: bool = False,
    ) -> tuple[Optional[Site], Optional[Client]]:
        """
        Resolve parsed header to Site + Client DB records.
        Returns (site, client) tuple. Either may be None if not found/created.
        """
        site = None
        client = None

        # If explicit override provided, use it
        if site_id_override:
            site = db.query(Site).filter(
                Site.site_id == site_id_override,
                Site.org_id == org_id,
            ).first()
            if site and site.client_id:
                client = db.query(Client).filter(Client.client_id == site.client_id).first()
            return site, client

        parsed_client_name = parsed_header.get("client", "")
        parsed_site_code = parsed_header.get("site_code", "")
        parsed_site_name = parsed_header.get("site_name", "")

        # Try matching client
        if client_id_override:
            client = db.query(Client).filter(
                Client.client_id == client_id_override,
                Client.org_id == org_id,
            ).first()
        elif parsed_client_name:
            # Match by client_code first
            client = db.query(Client).filter(
                Client.org_id == org_id,
                Client.client_code == parsed_client_name,
            ).first()
            # Then by client_name
            if not client:
                client = db.query(Client).filter(
                    Client.org_id == org_id,
                    Client.client_name.ilike(f"%{parsed_client_name}%"),
                ).first()

        # Try matching site
        if parsed_site_code:
            site = db.query(Site).filter(
                Site.org_id == org_id,
                Site.site_code == parsed_site_code,
            ).first()

        if not site and parsed_site_name:
            query = db.query(Site).filter(
                Site.org_id == org_id,
                Site.site_name.ilike(f"%{parsed_site_name}%"),
            )
            if client:
                query = query.filter(Site.client_id == client.client_id)
            site = query.first()

        # Auto-create if requested
        if auto_create and not client and parsed_client_name:
            client = Client(
                org_id=org_id,
                client_name=parsed_client_name,
                client_code=parsed_client_name,
                status="active",
            )
            db.add(client)
            db.flush()

        if auto_create and not site and parsed_site_name:
            site = Site(
                org_id=org_id,
                client_id=client.client_id if client else None,
                client_name=client.client_name if client else parsed_client_name,
                site_name=parsed_site_name,
                site_code=parsed_site_code,
                address=f"{parsed_site_name}, {parsed_header.get('province', 'South Africa')}",
            )
            db.add(site)
            db.flush()

        return site, client

    def import_roster(
        self,
        db: Session,
        file_content: bytes,
        filename: str,
        org_id: int,
        user_id: int,
        site_id_override: Optional[int] = None,
        client_id_override: Optional[int] = None,
        shift_start_time: str = "06:00",
        shift_end_time: str = "18:00",
        auto_create_site: bool = False,
        dry_run: bool = False,
    ) -> dict:
        """
        Main entry point. Parses Mafoko roster file and creates DB records.
        Set dry_run=True for preview mode (no DB writes).
        """
        file_hash = self.compute_file_hash(file_content)
        warnings = []
        errors = []

        # Check for duplicate upload
        if not dry_run:
            existing = db.query(RosterUploadLog).filter(
                RosterUploadLog.org_id == org_id,
                RosterUploadLog.file_hash == file_hash,
                RosterUploadLog.status.in_(["completed", "processing"]),
            ).first()
            if existing:
                return {
                    "status": "duplicate",
                    "message": f"This file was already uploaded (upload_id={existing.upload_id}, uploaded_at={existing.uploaded_at})",
                    "existing_upload_id": existing.upload_id,
                    "existing_roster_id": existing.roster_id,
                }

        # Load workbook
        try:
            wb = load_workbook(io.BytesIO(file_content), data_only=True)
        except Exception as e:
            return {"status": "error", "message": f"Could not read Excel file: {str(e)}"}

        # Find the right sheet
        ws = None
        for sheet_name in ["Site Roster", "Roster", "Sheet1"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                break
        if ws is None:
            ws = wb.active

        # Validate format
        if not self.detect_mafoko_format(ws):
            return {
                "status": "error",
                "message": "This file does not appear to be in Mafoko site roster format. "
                           "Expected 'Pers No' column and Mafoko header.",
            }

        # Parse header
        header_info = self.parse_header(ws)
        if not header_info.get("raw_header"):
            warnings.append("Could not parse Mafoko header. Site/client matching may fail.")

        # Parse date range — try filename first, then "Scheduled Period:" row in sheet
        start_date, end_date = self.parse_date_range(filename)
        if not start_date or not end_date:
            start_date, end_date = self.parse_date_range_from_sheet(ws)
        if not start_date or not end_date:
            warnings.append(f"Could not parse date range from filename or sheet. Using column numbers as day offsets from today.")
            start_date = date.today()
            end_date = start_date + timedelta(days=30)

        # Find data structure
        header_row, column_map = self.find_data_start(ws)
        if header_row is None:
            return {
                "status": "error",
                "message": "Could not find data header row (looking for 'Pers No' in column A).",
            }

        # Parse employee rows
        roster_data = self.parse_roster_data(ws, header_row, column_map)
        if not roster_data:
            return {
                "status": "error",
                "message": "No employee data found in the file.",
            }

        # Match employees
        employee_map, unmatched = self.match_employees(db, org_id, roster_data)
        for um in unmatched:
            warnings.append(f"Employee not found: {um['pers_no']} ({um['name']}). Skipping assignments.")

        # Resolve site
        site, client = self.resolve_site(
            db, org_id, header_info,
            site_id_override=site_id_override,
            client_id_override=client_id_override,
            auto_create=auto_create_site and not dry_run,
        )

        if not site and not dry_run and not auto_create_site:
            warnings.append(
                f"No matching site found for '{header_info.get('site_name', 'unknown')}'. "
                f"Provide site_id or enable auto_create_site."
            )

        # Generate suggestions
        from app.services.roster_suggestion_service import RosterSuggestionService

        suggestions = RosterSuggestionService.analyze(
            roster_data=roster_data,
            employee_map=employee_map,
            start_date=start_date,
            end_date=end_date,
            site=site,
            db=db,
            org_id=org_id,
        )

        # Calculate summary stats
        total_working_days = sum(len(r["working_days"]) for r in roster_data)
        matched_working_days = sum(
            len(r["working_days"]) for r in roster_data if r["pers_no"] in employee_map
        )

        # Build preview result
        result = {
            "status": "preview" if dry_run else "success",
            "parsed_header": {
                "province": header_info.get("province"),
                "client": header_info.get("client"),
                "site_code": header_info.get("site_code"),
                "site_name": header_info.get("site_name"),
                "schedule": header_info.get("schedule"),
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
            },
            "site_resolved": {
                "site_id": site.site_id if site else None,
                "site_name": site.site_name if site else None,
                "client_id": client.client_id if client else None,
                "client_name": client.client_name if client else None,
            } if site or client else None,
            "summary": {
                "employees_in_file": len(roster_data),
                "employees_matched": len(employee_map),
                "employees_unmatched": len(unmatched),
                "total_working_days": total_working_days,
                "matched_working_days": matched_working_days,
            },
            "unmatched_employees": unmatched,
            "suggestions": suggestions,
            "warnings": warnings[:50],
            "errors": errors[:50],
        }

        if dry_run:
            return result

        # ── Create DB records ────────────────────────────────────────────────
        if not site:
            return {
                **result,
                "status": "error",
                "message": "Cannot create roster without a resolved site. Provide site_id or enable auto_create_site.",
            }

        now = datetime.utcnow()
        roster_code = f"R-MAFOKO-{now.strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"

        # Create Roster record
        roster = Roster(
            org_id=org_id,
            roster_code=roster_code,
            name=f"Mafoko {header_info.get('site_name', 'Site')} {start_date.strftime('%d %b')} - {end_date.strftime('%d %b %Y')}",
            start_date=datetime.combine(start_date, time.min),
            end_date=datetime.combine(end_date, time.max),
            client_id=client.client_id if client else None,
            status="published",
            total_shifts=0,
            assigned_shifts=0,
            unassigned_shifts=0,
            total_cost=0.0,
            regular_pay_cost=0.0,
            overtime_cost=0.0,
            premium_cost=0.0,
            algorithm_used="mafoko_upload",
            created_by=user_id,
            published_at=now,
            published_by=user_id,
        )
        db.add(roster)
        db.flush()

        # Parse shift times
        try:
            s_start = time.fromisoformat(shift_start_time)
            s_end = time.fromisoformat(shift_end_time)
        except ValueError:
            s_start = time(6, 0)
            s_end = time(18, 0)

        # Group assignments by date to create shared Shifts
        # date -> list of (employee, row_data)
        date_assignments: dict[date, list] = {}
        for row in roster_data:
            employee = employee_map.get(row["pers_no"])
            if not employee:
                continue
            for day_offset in row["working_days"]:
                actual_date = start_date + timedelta(days=day_offset)
                if actual_date > end_date:
                    continue
                if actual_date not in date_assignments:
                    date_assignments[actual_date] = []
                date_assignments[actual_date].append((employee, row))

        # Create Shifts and ShiftAssignments
        shifts_created = 0
        assignments_created = 0
        total_cost = 0.0
        regular_cost = 0.0
        overtime_cost = 0.0
        premium_cost = 0.0

        for shift_date in sorted(date_assignments.keys()):
            emp_list = date_assignments[shift_date]
            required_staff = len(emp_list)

            shift_start_dt = datetime.combine(shift_date, s_start)
            # Handle overnight shifts
            if s_end <= s_start:
                shift_end_dt = datetime.combine(shift_date + timedelta(days=1), s_end)
            else:
                shift_end_dt = datetime.combine(shift_date, s_end)

            # Check for existing shift at this site+time
            existing_shift = db.query(Shift).filter(
                Shift.site_id == site.site_id,
                Shift.org_id == org_id,
                Shift.start_time == shift_start_dt,
                Shift.end_time == shift_end_dt,
            ).first()

            if existing_shift:
                shift = existing_shift
                shift.required_staff = max(shift.required_staff, required_staff)
            else:
                shift = Shift(
                    org_id=org_id,
                    site_id=site.site_id,
                    start_time=shift_start_dt,
                    end_time=shift_end_dt,
                    required_staff=required_staff,
                    status=ShiftStatus.CONFIRMED,
                )
                db.add(shift)
                db.flush()
                shifts_created += 1

            for employee, row_data in emp_list:
                # Check for existing assignment (idempotency)
                existing_assignment = db.query(ShiftAssignment).filter(
                    ShiftAssignment.shift_id == shift.shift_id,
                    ShiftAssignment.employee_id == employee.employee_id,
                ).first()
                if existing_assignment:
                    continue

                assignment = ShiftAssignment(
                    shift_id=shift.shift_id,
                    employee_id=employee.employee_id,
                    roster_id=roster.roster_id,
                    status=AssignmentStatus.CONFIRMED.value,
                    is_confirmed=True,
                    confirmation_datetime=now,
                    assigned_by=user_id,
                )

                try:
                    assignment.calculate_cost(shift, employee)
                except Exception as e:
                    logger.warning(f"Cost calculation failed for {row_data['pers_no']}: {e}")
                    duration = (shift_end_dt - shift_start_dt).total_seconds() / 3600
                    assignment.regular_hours = duration
                    assignment.regular_pay = duration * (employee.hourly_rate or 0)
                    assignment.total_cost = assignment.regular_pay

                total_cost += assignment.total_cost
                regular_cost += assignment.regular_pay
                overtime_cost += assignment.overtime_pay
                premium_cost += assignment.sunday_premium + assignment.holiday_premium + assignment.night_premium

                db.add(assignment)
                assignments_created += 1

        # Update roster stats
        roster.total_shifts = shifts_created
        roster.assigned_shifts = assignments_created
        roster.unassigned_shifts = max(0, shifts_created - assignments_created)
        roster.total_cost = total_cost
        roster.regular_pay_cost = regular_cost
        roster.overtime_cost = overtime_cost
        roster.premium_cost = premium_cost

        # Create upload log
        upload_log = RosterUploadLog(
            org_id=org_id,
            roster_id=roster.roster_id,
            filename=filename,
            file_hash=file_hash,
            upload_format="mafoko",
            uploaded_by=user_id,
            uploaded_at=now,
            parsed_province=header_info.get("province"),
            parsed_client=header_info.get("client"),
            parsed_site_code=header_info.get("site_code"),
            parsed_site_name=header_info.get("site_name"),
            parsed_schedule=header_info.get("schedule"),
            parsed_start_date=start_date,
            parsed_end_date=end_date,
            client_id=client.client_id if client else None,
            site_id=site.site_id,
            status="completed",
            employees_matched=len(employee_map),
            employees_unmatched=len(unmatched),
            shifts_created=shifts_created,
            assignments_created=assignments_created,
            total_cost=total_cost,
            warnings=warnings[:50] if warnings else None,
            errors=errors[:50] if errors else None,
            suggestions=suggestions[:50] if suggestions else None,
        )
        db.add(upload_log)

        try:
            db.commit()
        except Exception as e:
            db.rollback()
            logger.error(f"Mafoko roster upload commit failed: {e}")
            return {"status": "error", "message": f"Failed to save roster: {str(e)}"}

        result.update({
            "status": "success",
            "upload_id": upload_log.upload_id,
            "roster_id": roster.roster_id,
            "roster_code": roster.roster_code,
            "summary": {
                **result["summary"],
                "shifts_created": shifts_created,
                "assignments_created": assignments_created,
                "total_cost": round(total_cost, 2),
                "regular_cost": round(regular_cost, 2),
                "overtime_cost": round(overtime_cost, 2),
                "premium_cost": round(premium_cost, 2),
                "period": f"{start_date} to {end_date}",
            },
        })

        return result
