"""
Generic Roster Parser — auto-detects Excel roster formats and maps columns.

Three-tier detection:
1. Known format detection (Mafoko D/o grid, RostraCore template)
2. Heuristic column mapping via keyword matching
3. Fallback to user-provided column mapping
"""

import io
import logging
import re
from datetime import date, datetime, time, timedelta
from typing import Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# ── Column keyword groups for heuristic matching ─────────────────────────────

EMPLOYEE_NAME_KEYWORDS = [
    "employee name", "guard name", "guard", "name", "surname", "staff name",
    "full name", "worker", "officer", "security officer", "personnel",
    "employee", "staff",
]

EMPLOYEE_ID_KEYWORDS = [
    "employee id", "employee no", "emp id", "emp no", "emp number",
    "staff id", "staff no", "personnel no", "pers no", "id number",
    "employee_id", "employee_number", "worker id",
]

DATE_KEYWORDS = [
    "date", "shift date", "roster date", "work date", "day", "schedule date",
]

START_TIME_KEYWORDS = [
    "start time", "shift start", "time in", "start", "from", "clock in",
    "begin", "report time", "arrival",
]

END_TIME_KEYWORDS = [
    "end time", "shift end", "time out", "end", "to", "clock out",
    "finish", "departure", "knock off",
]

SITE_KEYWORDS = [
    "site", "site name", "location", "branch", "post", "station",
    "venue", "premises", "deployment",
]

CLIENT_KEYWORDS = [
    "client", "client name", "company", "customer",
]

GRADE_KEYWORDS = [
    "grade", "psira grade", "qualification", "level", "rank",
]

PSIRA_KEYWORDS = [
    "psira", "psira number", "psira no", "registration", "psira reg",
]

NOTES_KEYWORDS = [
    "notes", "comments", "remarks", "description",
]


# ── Keyword groups mapped to field names ─────────────────────────────────────

FIELD_KEYWORDS = {
    "employee_name": EMPLOYEE_NAME_KEYWORDS,
    "employee_id": EMPLOYEE_ID_KEYWORDS,
    "date": DATE_KEYWORDS,
    "start_time": START_TIME_KEYWORDS,
    "end_time": END_TIME_KEYWORDS,
    "site_name": SITE_KEYWORDS,
    "client_name": CLIENT_KEYWORDS,
    "grade": GRADE_KEYWORDS,
    "psira_number": PSIRA_KEYWORDS,
    "notes": NOTES_KEYWORDS,
}

# Fields that must be mapped for a valid import
REQUIRED_FIELDS = {"employee_name", "date"}

# RostraCore template expected headers
ROSTRACORE_HEADERS = {"date", "day", "shift start", "shift end", "guard name", "employee id", "psira number", "grade", "notes"}


class FormatDetection:
    """Result of format detection analysis."""

    def __init__(
        self,
        detected_format: str,
        confidence: float,
        headers: list[str],
        sample_rows: list[list],
        suggested_mapping: dict,
        header_row: int = 1,
        data_start_row: int = 2,
    ):
        self.detected_format = detected_format
        self.confidence = confidence
        self.headers = headers
        self.sample_rows = sample_rows
        self.suggested_mapping = suggested_mapping
        self.header_row = header_row
        self.data_start_row = data_start_row

    def to_dict(self) -> dict:
        return {
            "detected_format": self.detected_format,
            "confidence": self.confidence,
            "headers": self.headers,
            "sample_rows": self.sample_rows,
            "suggested_mapping": self.suggested_mapping,
            "header_row": self.header_row,
            "data_start_row": self.data_start_row,
        }


class RosterEntry:
    """A single parsed roster entry (one employee-shift assignment)."""

    def __init__(
        self,
        employee_identifier: str,
        employee_id: Optional[str],
        shift_date: date,
        start_time: Optional[time],
        end_time: Optional[time],
        site_name: Optional[str] = None,
        client_name: Optional[str] = None,
        grade: Optional[str] = None,
        psira_number: Optional[str] = None,
        notes: Optional[str] = None,
        row_number: int = 0,
    ):
        self.employee_identifier = employee_identifier
        self.employee_id = employee_id
        self.shift_date = shift_date
        self.start_time = start_time
        self.end_time = end_time
        self.site_name = site_name
        self.client_name = client_name
        self.grade = grade
        self.psira_number = psira_number
        self.notes = notes
        self.row_number = row_number

    def to_dict(self) -> dict:
        return {
            "employee_identifier": self.employee_identifier,
            "employee_id": self.employee_id,
            "date": self.shift_date.isoformat(),
            "start_time": self.start_time.strftime("%H:%M") if self.start_time else None,
            "end_time": self.end_time.strftime("%H:%M") if self.end_time else None,
            "site_name": self.site_name,
            "client_name": self.client_name,
            "grade": self.grade,
            "psira_number": self.psira_number,
            "notes": self.notes,
            "row_number": self.row_number,
        }


class ParseResult:
    """Result of parsing a roster file."""

    def __init__(
        self,
        entries: list[RosterEntry],
        date_range: tuple[Optional[date], Optional[date]],
        warnings: list[str],
        errors: list[str],
        detected_format: str,
        site_name: Optional[str] = None,
        client_name: Optional[str] = None,
    ):
        self.entries = entries
        self.date_range = date_range
        self.warnings = warnings
        self.errors = errors
        self.detected_format = detected_format
        self.site_name = site_name
        self.client_name = client_name


class GenericRosterParser:
    """Smart Excel roster parser with auto-detection and column mapping."""

    def detect_format(self, file_bytes: bytes, filename: str) -> FormatDetection:
        """
        Analyze an Excel file and detect its roster format.
        Returns format type, confidence, headers, sample rows, and suggested column mapping.
        """
        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            raise ValueError(f"Could not read Excel file: {e}")

        ws = self._find_roster_sheet(wb)

        # Tier 1: Check for Mafoko D/o grid format
        if self._is_mafoko_format(ws):
            return FormatDetection(
                detected_format="mafoko_do_grid",
                confidence=0.95,
                headers=["Mafoko D/o Grid Format Detected"],
                sample_rows=[],
                suggested_mapping={},
                header_row=0,
                data_start_row=0,
            )

        # Tier 1: Check for RostraCore template format
        header_row, headers = self._find_header_row(ws)
        if header_row and self._is_rostracore_template(headers):
            return FormatDetection(
                detected_format="rostracore_template",
                confidence=0.95,
                headers=headers,
                sample_rows=self._get_sample_rows(ws, header_row, len(headers)),
                suggested_mapping=self._map_rostracore_template(headers),
                header_row=header_row,
                data_start_row=header_row + 1,
            )

        # Tier 2: Heuristic column mapping
        if header_row and headers:
            mapping, confidence = self._heuristic_mapping(headers)
            return FormatDetection(
                detected_format="generic" if confidence >= 0.5 else "unknown",
                confidence=confidence,
                headers=headers,
                sample_rows=self._get_sample_rows(ws, header_row, len(headers)),
                suggested_mapping=mapping,
                header_row=header_row,
                data_start_row=header_row + 1,
            )

        # Tier 3: Could not detect — return raw data for user mapping
        headers, header_row = self._extract_any_headers(ws)
        return FormatDetection(
            detected_format="unknown",
            confidence=0.0,
            headers=headers,
            sample_rows=self._get_sample_rows(ws, header_row, len(headers)) if headers else [],
            suggested_mapping={},
            header_row=header_row or 1,
            data_start_row=(header_row or 1) + 1,
        )

    def parse(
        self,
        file_bytes: bytes,
        filename: str,
        column_mapping: dict,
        default_start_time: str = "06:00",
        default_end_time: str = "18:00",
        default_site_name: Optional[str] = None,
        default_client_name: Optional[str] = None,
    ) -> ParseResult:
        """
        Parse a roster Excel file using the provided column mapping.
        Returns a list of RosterEntry objects.
        """
        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            raise ValueError(f"Could not read Excel file: {e}")

        ws = self._find_roster_sheet(wb)
        entries = []
        warnings = []
        errors = []

        # Determine header and data start rows
        header_row_num = column_mapping.get("_header_row")
        if not header_row_num:
            header_row_num, _ = self._find_header_row(ws)
            if not header_row_num:
                header_row_num = 1

        data_start = column_mapping.get("_data_start_row", header_row_num + 1)

        # Parse default times
        try:
            def_start = datetime.strptime(default_start_time, "%H:%M").time()
        except ValueError:
            def_start = time(6, 0)
        try:
            def_end = datetime.strptime(default_end_time, "%H:%M").time()
        except ValueError:
            def_end = time(18, 0)

        # Extract column indices from mapping (0-indexed in mapping, 1-indexed in openpyxl)
        name_col = column_mapping.get("employee_name")
        id_col = column_mapping.get("employee_id")
        date_col = column_mapping.get("date")
        start_col = column_mapping.get("start_time")
        end_col = column_mapping.get("end_time")
        site_col = column_mapping.get("site_name")
        client_col = column_mapping.get("client_name")
        grade_col = column_mapping.get("grade")
        psira_col = column_mapping.get("psira_number")
        notes_col = column_mapping.get("notes")

        # We need at least employee identifier and date
        if name_col is None and id_col is None:
            errors.append("Column mapping must include at least employee_name or employee_id.")
            return ParseResult([], (None, None), warnings, errors, "generic")

        if date_col is None:
            errors.append("Column mapping must include date.")
            return ParseResult([], (None, None), warnings, errors, "generic")

        all_dates = []

        for r in range(data_start, ws.max_row + 1):
            # Read employee identifier
            emp_name = self._cell_str(ws, r, name_col) if name_col is not None else ""
            emp_id = self._cell_str(ws, r, id_col) if id_col is not None else ""

            if not emp_name and not emp_id:
                continue  # Skip empty rows

            # Read date
            shift_date = self._parse_date(ws, r, date_col)
            if not shift_date:
                raw_date = self._cell_str(ws, r, date_col)
                if raw_date:
                    errors.append(f"Row {r}: Could not parse date '{raw_date}'")
                continue

            all_dates.append(shift_date)

            # Read times (optional — use defaults if not mapped)
            start_t = self._parse_time(ws, r, start_col) if start_col is not None else def_start
            end_t = self._parse_time(ws, r, end_col) if end_col is not None else def_end

            # Read optional fields
            site_name = self._cell_str(ws, r, site_col) if site_col is not None else default_site_name
            client_name = self._cell_str(ws, r, client_col) if client_col is not None else default_client_name
            grade = self._cell_str(ws, r, grade_col) if grade_col is not None else None
            psira = self._cell_str(ws, r, psira_col) if psira_col is not None else None
            notes = self._cell_str(ws, r, notes_col) if notes_col is not None else None

            entries.append(RosterEntry(
                employee_identifier=emp_name or emp_id,
                employee_id=emp_id or None,
                shift_date=shift_date,
                start_time=start_t,
                end_time=end_t,
                site_name=site_name,
                client_name=client_name,
                grade=grade,
                psira_number=psira,
                notes=notes,
                row_number=r,
            ))

        date_range = (min(all_dates), max(all_dates)) if all_dates else (None, None)

        # Try to extract site/client from spreadsheet metadata
        detected_site = default_site_name
        detected_client = default_client_name
        if not detected_site or not detected_client:
            meta_site, meta_client = self._extract_metadata(ws, header_row_num)
            if not detected_site and meta_site:
                detected_site = meta_site
            if not detected_client and meta_client:
                detected_client = meta_client

        return ParseResult(
            entries=entries,
            date_range=date_range,
            warnings=warnings[:50],
            errors=errors[:50],
            detected_format="generic",
            site_name=detected_site,
            client_name=detected_client,
        )

    # ── Private helpers ──────────────────────────────────────────────────────

    def _find_roster_sheet(self, wb):
        """Find the most likely roster worksheet."""
        for name in ["Site Roster", "Roster", "Schedule", "Shifts", "Sheet1"]:
            if name in wb.sheetnames:
                return wb[name]
        return wb.active

    def _is_mafoko_format(self, ws) -> bool:
        """Check if worksheet matches Mafoko D/o grid format."""
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(row=r, column=c).value
                if val and "mafoko" in str(val).lower() and "roster" in str(val).lower():
                    return True
        for r in range(1, min(15, ws.max_row + 1)):
            val = ws.cell(row=r, column=1).value
            if val and str(val).strip().lower() in ("pers no", "pers no.", "personnel no"):
                # Check for D/o grid columns (day numbers 1-31)
                for c in range(6, min(40, ws.max_column + 1)):
                    col_val = ws.cell(row=r, column=c).value
                    if col_val:
                        try:
                            d = int(str(col_val).strip())
                            if 1 <= d <= 31:
                                return True
                        except (ValueError, TypeError):
                            pass
        return False

    def _is_rostracore_template(self, headers: list[str]) -> bool:
        """Check if headers match RostraCore template format."""
        header_set = {h.lower().strip() for h in headers if h}
        # Must have at least 5 of the 9 RostraCore headers
        matches = header_set & ROSTRACORE_HEADERS
        return len(matches) >= 5

    def _map_rostracore_template(self, headers: list[str]) -> dict:
        """Map RostraCore template headers to field indices."""
        mapping = {}
        for i, h in enumerate(headers):
            h_lower = (h or "").lower().strip()
            if h_lower == "guard name":
                mapping["employee_name"] = i
            elif h_lower == "employee id":
                mapping["employee_id"] = i
            elif h_lower == "date":
                mapping["date"] = i
            elif h_lower == "shift start":
                mapping["start_time"] = i
            elif h_lower == "shift end":
                mapping["end_time"] = i
            elif h_lower == "psira number":
                mapping["psira_number"] = i
            elif h_lower == "grade":
                mapping["grade"] = i
            elif h_lower == "notes":
                mapping["notes"] = i
        return mapping

    def _find_header_row(self, ws) -> tuple[Optional[int], list[str]]:
        """
        Find the header row by looking for a row with multiple text cells
        that look like column headers.
        """
        best_row = None
        best_headers = []
        best_score = 0

        for r in range(1, min(25, ws.max_row + 1)):
            row_values = []
            non_empty = 0
            for c in range(1, min(30, ws.max_column + 1)):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    row_values.append(str(val).strip())
                    non_empty += 1
                else:
                    row_values.append("")

            if non_empty < 2:
                continue

            # Score this row as a potential header
            score = 0
            for val in row_values:
                if not val:
                    continue
                val_lower = val.lower()
                # Check if this cell matches any known keyword
                for keywords in FIELD_KEYWORDS.values():
                    if val_lower in keywords or any(kw in val_lower for kw in keywords):
                        score += 2
                        break
                # Bonus for short text (headers are usually short)
                if len(val) < 30 and not val.replace(" ", "").isdigit():
                    score += 1

            if score > best_score:
                best_score = score
                best_row = r
                # Trim trailing empty cells
                while best_headers and not best_headers[-1]:
                    best_headers.pop()
                best_headers = [v for v in row_values if v or row_values.index(v) < non_empty]

        # Trim trailing empty from best_headers
        while best_headers and not best_headers[-1]:
            best_headers.pop()

        return best_row, best_headers

    def _extract_any_headers(self, ws) -> tuple[list[str], Optional[int]]:
        """Fallback: return the first non-empty row as headers."""
        for r in range(1, min(20, ws.max_row + 1)):
            vals = []
            non_empty = 0
            for c in range(1, min(30, ws.max_column + 1)):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    vals.append(str(val).strip())
                    non_empty += 1
                else:
                    vals.append("")
            if non_empty >= 2:
                while vals and not vals[-1]:
                    vals.pop()
                return vals, r
        return [], None

    def _heuristic_mapping(self, headers: list[str]) -> tuple[dict, float]:
        """
        Map column headers to field names using keyword matching.
        Returns (mapping, confidence).
        """
        mapping = {}
        matched_fields = set()

        for i, header in enumerate(headers):
            if not header:
                continue
            h_lower = header.lower().strip()

            best_field = None
            best_score = 0

            for field, keywords in FIELD_KEYWORDS.items():
                if field in matched_fields:
                    continue

                for kw in keywords:
                    # Exact match
                    if h_lower == kw:
                        score = 10
                    # Header contains keyword
                    elif kw in h_lower:
                        score = 5 + (len(kw) / len(h_lower)) * 3
                    # Keyword contains header (partial match)
                    elif h_lower in kw and len(h_lower) > 2:
                        score = 3
                    else:
                        continue

                    if score > best_score:
                        best_score = score
                        best_field = field

            if best_field and best_score >= 3:
                mapping[best_field] = i
                matched_fields.add(best_field)

        # Calculate confidence based on matched required + optional fields
        required_matched = sum(1 for f in REQUIRED_FIELDS if f in mapping)
        total_matched = len(mapping)

        if required_matched == 0:
            confidence = 0.0
        elif required_matched < len(REQUIRED_FIELDS):
            confidence = 0.3 + (total_matched * 0.05)
        else:
            confidence = 0.5 + min(total_matched * 0.08, 0.45)

        return mapping, min(confidence, 0.95)

    def _get_sample_rows(self, ws, header_row: int, num_cols: int, max_rows: int = 5) -> list[list]:
        """Get sample data rows after the header."""
        samples = []
        for r in range(header_row + 1, min(header_row + max_rows + 1, ws.max_row + 1)):
            row_data = []
            has_data = False
            for c in range(1, num_cols + 1):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    has_data = True
                    # Convert datetime to string for JSON serialization
                    if isinstance(val, datetime):
                        row_data.append(val.strftime("%Y-%m-%d %H:%M"))
                    elif isinstance(val, date):
                        row_data.append(val.isoformat())
                    else:
                        row_data.append(str(val))
                else:
                    row_data.append("")
            if has_data:
                samples.append(row_data)
        return samples

    def _cell_str(self, ws, row: int, col_idx: Optional[int]) -> str:
        """Read a cell value as a stripped string. col_idx is 0-indexed."""
        if col_idx is None:
            return ""
        val = ws.cell(row=row, column=col_idx + 1).value  # openpyxl is 1-indexed
        if val is None:
            return ""
        return str(val).strip()

    def _parse_date(self, ws, row: int, col_idx: Optional[int]) -> Optional[date]:
        """Parse a date from a cell. col_idx is 0-indexed."""
        if col_idx is None:
            return None
        val = ws.cell(row=row, column=col_idx + 1).value
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        # Try string parsing
        val_str = str(val).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y"):
            try:
                return datetime.strptime(val_str, fmt).date()
            except ValueError:
                continue
        return None

    def _parse_time(self, ws, row: int, col_idx: Optional[int]) -> Optional[time]:
        """Parse a time from a cell. col_idx is 0-indexed."""
        if col_idx is None:
            return None
        val = ws.cell(row=row, column=col_idx + 1).value
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.time()
        if isinstance(val, time):
            return val
        val_str = str(val).strip()
        for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M%p"):
            try:
                return datetime.strptime(val_str, fmt).time()
            except ValueError:
                continue
        return None

    def _extract_metadata(self, ws, header_row: int) -> tuple[Optional[str], Optional[str]]:
        """Try to extract site/client name from metadata rows above the header."""
        site_name = None
        client_name = None
        for r in range(1, min(header_row, 10)):
            for c in range(1, 3):
                label = ws.cell(row=r, column=c).value
                if not label:
                    continue
                label_str = str(label).strip().lower()
                if "site" in label_str and ":" in label_str:
                    parts = str(label).split(":", 1)
                    if len(parts) > 1 and parts[1].strip():
                        site_name = parts[1].strip()
                elif "site" in label_str:
                    val = ws.cell(row=r, column=c + 1).value
                    if val:
                        site_name = str(val).strip()
                elif "client" in label_str and ":" in label_str:
                    parts = str(label).split(":", 1)
                    if len(parts) > 1 and parts[1].strip():
                        client_name = parts[1].strip()
                elif "client" in label_str:
                    val = ws.cell(row=r, column=c + 1).value
                    if val:
                        client_name = str(val).strip()
        return site_name, client_name
